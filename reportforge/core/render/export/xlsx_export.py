# export/xlsx_export.py — Excel export via openpyxl
from __future__ import annotations
from pathlib import Path
from typing import Any


def export_xlsx(data: dict, output: str | Path,
                dataset: str = "items",
                fields: list[str] = None,
                title: str = "Report",
                sheet_name: str = "Data") -> Path:
    """
    Export report data to a formatted Excel file.
    
    Supports:
      - Multiple sheets (one per named dataset)
      - Auto-column widths
      - Header row styling
      - Number formatting
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("XLSX export requires openpyxl: pip install openpyxl")

    p = Path(output)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Header style ──────────────────────────────────────────────
    hdr_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1A1A2E")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    hdr_border= Border(bottom=thin, right=thin)

    def _write_sheet(ws, rows, header_fields):
        # Header
        for ci, field in enumerate(header_fields, 1):
            cell = ws.cell(row=1, column=ci, value=field.upper())
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = hdr_border
        ws.row_dimensions[1].height = 20

        # Data rows
        for ri, row in enumerate(rows, 2):
            alt_fill = PatternFill("solid", fgColor="F4F4F2") if ri % 2 == 0 else None
            for ci, field in enumerate(header_fields, 1):
                val = row.get(field, "")
                cell = ws.cell(row=ri, column=ci, value=_coerce(val))
                if alt_fill:
                    cell.fill = alt_fill
                cell.alignment = Alignment(vertical="center")
                # Number formatting
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(val, int):
                    cell.number_format = "#,##0"

        # Auto-width
        for ci, field in enumerate(header_fields, 1):
            col_letter = get_column_letter(ci)
            max_len = max((len(str(r.get(field, ""))) for r in rows), default=0)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, len(field) + 2), 40)

        # Freeze header
        ws.freeze_panes = "A2"

    # ── Primary sheet ─────────────────────────────────────────────
    rows = data.get(dataset, [])
    if isinstance(rows, list) and rows:
        ws = wb.active
        ws.title = sheet_name[:31]
        if not fields:
            fields = list(rows[0].keys())
        _write_sheet(ws, rows, fields)

        # Report metadata sheet
        meta_ws = wb.create_sheet("Info")
        meta_ws["A1"] = "Report"
        meta_ws["B1"] = title
        meta_ws["A2"] = "Generated"
        import datetime
        meta_ws["B2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_ws["A3"] = "Total rows"
        meta_ws["B3"] = len(rows)
    else:
        # Write all list-type keys to separate sheets
        first = True
        for key, val in data.items():
            if not isinstance(val, list) or not val:
                continue
            if first:
                ws = wb.active
                ws.title = key[:31]
                first = False
            else:
                ws = wb.create_sheet(key[:31])
            flds = list(val[0].keys())
            _write_sheet(ws, val, flds)

    wb.save(str(p))
    return p


def _coerce(val: Any) -> Any:
    """Coerce value to Excel-safe type."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (int, float)):
        return val
    try:
        return float(val)
    except (TypeError, ValueError):
        pass
    return str(val)
