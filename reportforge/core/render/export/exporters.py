# core/render/export/exporters.py
# Export formats: PDF, HTML, PNG, CSV, XLSX
from __future__ import annotations
import csv, io, json
from pathlib import Path
from typing import Any


class Exporter:
    """
    Multi-format exporter for ReportForge.

    Supported formats:
        pdf   → WeasyPrint
        html  → plain file save
        png   → WeasyPrint image
        csv   → flat CSV of detail items
        xlsx  → openpyxl workbook
    """

    def __init__(self, layout_raw: dict, data: dict,
                 debug: bool = False):
        self._layout_raw = layout_raw
        self._data       = data
        self._debug      = debug
        self._html: str | None = None  # cached rendered HTML

    def render_html(self) -> str:
        if self._html is None:
            from ..engines.advanced_engine import AdvancedHtmlEngine
            self._html = AdvancedHtmlEngine(
                self._layout_raw, self._data, debug=self._debug
            ).render()
        return self._html

    # ── HTML ──────────────────────────────────────────────────────

    def to_html(self, output_path: str | Path | None = None) -> str:
        html = self.render_html()
        if output_path:
            Path(output_path).write_text(html, encoding="utf-8")
        return html

    # ── PDF ───────────────────────────────────────────────────────

    def to_pdf(self, output_path: str | Path) -> Path:
        from ..engines.pdf_generator import PdfGenerator
        html = self.render_html()
        p    = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        PdfGenerator().from_html_to_file(html, p)
        return p

    # ── PNG ───────────────────────────────────────────────────────

    def to_png(self, output_path: str | Path,
               resolution: int = 150) -> Path:
        """Render first page to PNG via WeasyPrint."""
        try:
            import weasyprint
        except ImportError:
            raise ImportError("PNG export requires WeasyPrint: pip install weasyprint")

        html  = self.render_html()
        p     = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        doc   = weasyprint.HTML(string=html)
        pages = doc.render()
        if not pages.pages:
            raise RuntimeError("No pages rendered")
        # Export first page
        surface, _w, _h = pages.pages[0].paint(
            target=None, left_x=0, top_y=0, right_x=0, bottom_y=0
        )
        # Use Cairo to write PNG
        try:
            surface.write_to_png(str(p))
        except AttributeError:
            # Alternative path via weasyprint's document
            from weasyprint import document as _doc
            pages.write_png(str(p), resolution=resolution)
        return p

    # ── DOCX ──────────────────────────────────────────────────────

    def to_docx(self, output_path: str | Path) -> Path:
        """Export to Microsoft Word (.docx) via python-docx."""
        from .docx_export import export_docx
        return export_docx(self._layout_raw, self._data, output_path)

    # ── RTF ───────────────────────────────────────────────────────

    def to_rtf(self, output_path: str | Path) -> Path:
        """Export to Rich Text Format (.rtf) — no external dependencies."""
        from .rtf_export import export_rtf
        return export_rtf(self._layout_raw, self._data, output_path)

    # ── CSV ───────────────────────────────────────────────────────

    def to_csv(self, output_path: str | Path | None = None,
               dataset: str = "items",
               encoding: str = "utf-8-sig",   # BOM for Excel
               delimiter: str = ",") -> str:
        """
        Export detail dataset to CSV.
        Columns are inferred from layout field elements in the 'det' section.
        """
        from ..pipeline.normalizer import normalize_layout
        from ..resolvers.field_resolver import FieldResolver, format_value

        norm     = normalize_layout(self._layout_raw)
        resolver = FieldResolver(self._data)

        # Get columns from detail section elements
        det_ids  = {s["id"] for s in norm["sections"] if s["stype"] == "det"}
        col_els  = [e for e in norm["elements"]
                    if e.get("sectionId") in det_ids
                    and e.get("type") in ("field", "text")
                    and e.get("fieldPath")]

        items = self._data.get(dataset, [])

        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=delimiter)

        # Header row
        headers = []
        for e in col_els:
            label = e.get("content") or e.get("fieldPath", "")
            headers.append(label)
        writer.writerow(headers)

        # Data rows
        for item in items:
            row = []
            res = resolver.with_item(item)
            for e in col_els:
                fp  = e.get("fieldPath", "")
                fmt = e.get("fieldFmt")
                try:
                    val = res.get(fp, "")
                    row.append(format_value(val, fmt) if fmt else (str(val) if val != "" else ""))
                except Exception:
                    row.append("")
            writer.writerow(row)

        csv_str = buf.getvalue()

        if output_path:
            p = Path(output_path)
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(csv_str, encoding=encoding)

        return csv_str

    # ── XLSX ──────────────────────────────────────────────────────

    def to_xlsx(self, output_path: str | Path,
                dataset: str = "items",
                sheet_name: str = "Report") -> Path:
        """Export to Excel workbook with header, data, and auto-width columns."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("XLSX export requires openpyxl: pip install openpyxl")

        from ..pipeline.normalizer import normalize_layout
        from ..resolvers.field_resolver import FieldResolver, format_value

        norm     = normalize_layout(self._layout_raw)
        resolver = FieldResolver(self._data)

        det_ids  = {s["id"] for s in norm["sections"] if s["stype"] == "det"}
        col_els  = [e for e in norm["elements"]
                    if e.get("sectionId") in det_ids
                    and e.get("type") in ("field", "text")
                    and e.get("fieldPath")]
        col_els.sort(key=lambda e: e.get("x", 0))

        items    = self._data.get(dataset, [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        hdr_font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1A3A6B")
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin      = Side(style="thin", color="CCCCCC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill  = PatternFill("solid", fgColor="F4F4F2")

        # Report title (row 1)
        report_name = norm.get("name", "Report")
        ws.cell(1, 1, report_name).font = Font(name="Arial", bold=True, size=12, color="1A3A6B")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(1, len(col_els)))

        # Header row (row 2)
        for ci, e in enumerate(col_els, 1):
            label = e.get("content") or _field_label(e.get("fieldPath",""))
            cell  = ws.cell(2, ci, label)
            cell.font   = hdr_font
            cell.fill   = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

        # Data rows
        col_widths = [len(_field_label(e.get("fieldPath",""))) + 2 for e in col_els]
        for ri, item in enumerate(items):
            excel_row = ri + 3
            res = resolver.with_item(item)
            for ci, e in enumerate(col_els, 1):
                fp  = e.get("fieldPath", "")
                fmt = e.get("fieldFmt")
                align_h = e.get("align", "left")
                try:
                    raw = res.get(fp, "")
                    if fmt:
                        val = format_value(raw, fmt)
                    else:
                        # Keep numeric for Excel
                        try: val = float(raw) if str(raw).replace(".", "").replace("-","").isdigit() else raw
                        except: val = raw
                except Exception:
                    val = ""

                cell = ws.cell(excel_row, ci, val)
                cell.border    = border
                cell.alignment = Alignment(horizontal=align_h, vertical="center")
                if ri % 2 == 1:
                    cell.fill = alt_fill

                col_widths[ci-1] = max(col_widths[ci-1], len(str(val)) + 2)

        # Auto column widths (max 40)
        for ci, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = min(width, 40)

        # Freeze header
        ws.freeze_panes = "A3"

        # Auto-filter on data
        if col_els and items:
            ws.auto_filter.ref = (
                f"A2:{get_column_letter(len(col_els))}{len(items)+2}"
            )

        p = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return p


def _field_label(fp: str) -> str:
    """Convert field path to human label: item.unit_price → Unit Price"""
    part = fp.split(".")[-1] if "." in fp else fp
    return part.replace("_", " ").replace("-"," ").title()


# ── Convenience functions ─────────────────────────────────────────

def export(layout_raw: dict, data: dict, output_path: str | Path,
           fmt: str | None = None, **kwargs) -> Path | str:
    """
    One-call export.

    fmt is inferred from output_path extension if not given.
    Returns output path (or HTML string for fmt='html').
    """
    p = Path(output_path)
    if fmt is None:
        fmt = p.suffix.lstrip(".").lower() or "html"

    ex = Exporter(layout_raw, data, **{k: v for k, v in kwargs.items()
                                        if k in ("debug",)})
    if fmt == "html":  return ex.to_html(p)
    if fmt == "pdf":   return ex.to_pdf(p)
    if fmt == "png":   return ex.to_png(p)
    if fmt == "csv":   return ex.to_csv(p)
    if fmt == "xlsx":  return ex.to_xlsx(p)
    if fmt == "docx":  return ex.to_docx(p)
    if fmt == "rtf":   return ex.to_rtf(p)
    raise ValueError(f"Unsupported export format: {fmt!r}")
