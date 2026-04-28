from __future__ import annotations

import csv
import io
from pathlib import Path

from ..pipeline.normalizer import normalize_layout
from ..resolvers.field_resolver import FieldResolver, format_value


def _field_label(fp: str) -> str:
    part = fp.split(".")[-1] if "." in fp else fp
    return part.replace("_", " ").replace("-", " ").title()


def to_csv(exporter, output_path: str | Path | None = None, dataset: str = "items", encoding: str = "utf-8-sig", delimiter: str = ",") -> str:
    norm = normalize_layout(exporter._layout_raw)
    resolver = FieldResolver(exporter._data)
    det_ids = {s["id"] for s in norm["sections"] if s["stype"] == "det"}
    col_els = [e for e in norm["elements"] if e.get("sectionId") in det_ids and e.get("type") in ("field", "text") and e.get("fieldPath")]
    items = exporter._data.get(dataset, [])
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerow([e.get("content") or e.get("fieldPath", "") for e in col_els])
    for item in items:
        row = []
        res = resolver.with_item(item)
        for e in col_els:
            fp = e.get("fieldPath", "")
            fmt = e.get("fieldFmt")
            try:
                val = res.get(fp, "")
                row.append(format_value(val, fmt) if fmt else (str(val) if val != "" else ""))
            except Exception:
                row.append("")
        writer.writerow(row)
    csv_str = buf.getvalue()
    if output_path:
        p = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(csv_str, encoding=encoding)
    return csv_str


def to_xlsx(exporter, output_path: str | Path, dataset: str = "items", sheet_name: str = "Report") -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("XLSX export requires openpyxl: pip install openpyxl")

    norm = normalize_layout(exporter._layout_raw)
    resolver = FieldResolver(exporter._data)
    det_ids = {s["id"] for s in norm["sections"] if s["stype"] == "det"}
    col_els = [e for e in norm["elements"] if e.get("sectionId") in det_ids and e.get("type") in ("field", "text") and e.get("fieldPath")]
    col_els.sort(key=lambda e: e.get("x", 0))
    items = exporter._data.get(dataset, [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A3A6B")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F4F4F2")

    report_name = norm.get("name", "Report")
    ws.cell(1, 1, report_name).font = Font(name="Arial", bold=True, size=12, color="1A3A6B")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(col_els)))

    for ci, e in enumerate(col_els, 1):
        label = e.get("content") or _field_label(e.get("fieldPath", ""))
        cell = ws.cell(2, ci, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    col_widths = [len(_field_label(e.get("fieldPath", ""))) + 2 for e in col_els]
    for ri, item in enumerate(items):
        excel_row = ri + 3
        res = resolver.with_item(item)
        for ci, e in enumerate(col_els, 1):
            fp = e.get("fieldPath", "")
            fmt = e.get("fieldFmt")
            align_h = e.get("align", "left")
            try:
                raw = res.get(fp, "")
                if fmt:
                    val = format_value(raw, fmt)
                else:
                    try:
                        val = float(raw) if str(raw).replace(".", "").replace("-", "").isdigit() else raw
                    except Exception:
                        val = raw
            except Exception:
                val = ""
            cell = ws.cell(excel_row, ci, val)
            cell.border = border
            cell.alignment = Alignment(horizontal=align_h, vertical="center")
            if ri % 2 == 1:
                cell.fill = alt_fill
            col_widths[ci - 1] = max(col_widths[ci - 1], len(str(val)) + 2)

    for ci, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 40)
    ws.freeze_panes = "A3"
    if col_els and items:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(col_els))}{len(items)+2}"

    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
    return p
